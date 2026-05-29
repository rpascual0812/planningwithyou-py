"""Build Excel exports for invitation RSVP submissions."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def build_rsvp_xlsx(
    *,
    rows: list[dict],
    columns: list[dict],
    sheet_title: str = 'RSVPs',
) -> bytes:
    """Return .xlsx bytes for RSVP rows with dynamic field columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or 'RSVPs'

    headers = ['Submitted At'] + [str(col.get('label') or col.get('id') or '') for col in columns]
    ws.append(headers)

    for row in rows:
        fields_data = row.get('fields_data') or {}
        if not isinstance(fields_data, dict):
            fields_data = {}
        created_at = row.get('created_at') or ''
        ws.append(
            [created_at, *[str(fields_data.get(col.get('id') or '', '') or '') for col in columns]],
        )

    for idx, header in enumerate(headers, start=1):
        width = max(len(header), 12)
        ws.column_dimensions[get_column_letter(idx)].width = min(width + 2, 48)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
